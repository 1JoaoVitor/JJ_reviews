import os
import requests # type: ignore
import unicodedata
from dotenv import load_dotenv  # type: ignore
from openpyxl import load_workbook  # type: ignore

# --- CONFIGURAÇÕES ---
load_dotenv()
TMDB_API_KEY = os.getenv("VITE_TMDB_API_KEY")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
EXCEL_FILE = "data-processor/filmes.xlsx"
# Aumentei a tolerância para 2 anos (ajuda em filmes de fim de ano ou festivais)
YEAR_TOLERANCE = 2 

def normalize_txt(text):
    """Remove acentos e deixa minúsculo para comparação"""
    if not text: return ""
    return unicodedata.normalize('NFKD', str(text)).encode('ASCII', 'ignore').decode('utf-8').lower().strip()

def get_movie_director(tmdb_id):
    """Busca os diretores na API"""
    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}/credits?api_key={TMDB_API_KEY}"
    try:
        res = requests.get(url).json()
        crew = res.get('crew', [])
        directors = [m['name'] for m in crew if m['job'] == 'Director']
        return directors
    except:
        return []

def check_director_match(tmdb_directors, excel_director):
    """Verifica match parcial de nomes (Ex: 'Del Toro' bate com 'Guillermo del Toro')"""
    if not excel_director or not tmdb_directors: return False
    
    target_tokens = set(normalize_txt(excel_director).split())
    
    for director in tmdb_directors:
        curr_tokens = set(normalize_txt(director).split())
        # Se houver interseção significativa de nomes (ex: sobrenome bate)
        if target_tokens & curr_tokens:
            return True
    return False

def check_year_match(tmdb_date, target_year):
    if not target_year or not tmdb_date: return True # Se não tem data, aceita
    try:
        tmdb_year = int(tmdb_date.split('-')[0])
        target_year = int(target_year)
        return abs(tmdb_year - target_year) <= YEAR_TOLERANCE
    except:
        return True # Se der erro no parse, aceita na dúvida

def find_best_match(nome, ano, diretor_excel):
    if not nome: return None

    # Busca na API
    url = f"https://api.themoviedb.org/3/search/movie?api_key={TMDB_API_KEY}&query={nome}&language=pt-BR"
    if ano: url += f"&year={ano}"
    
    try:
        results = requests.get(url).json().get('results', [])
    except:
        return None

    # Fallback: Se não achou com ano, tenta SEM ano (para casos como Frankstein 2025 que a API pode ter data diferente)
    if not results and ano:
        url_no_year = f"https://api.themoviedb.org/3/search/movie?api_key={TMDB_API_KEY}&query={nome}&language=pt-BR"
        results = requests.get(url_no_year).json().get('results', [])
    
    if not results: return None

    # --- LÓGICA DE VALIDAÇÃO ---
    print(f"    (Validando {len(results[:5])} candidatos na API...)", end="\r")
    
    for movie in results[:5]: # Olha apenas os top 5 resultados
        
        # 1. Checagem de DIRETOR (Prioridade Máxima)
        if diretor_excel:
            tmdb_directors = get_movie_director(movie['id'])
            if check_director_match(tmdb_directors, diretor_excel):
                # Achamos o diretor! Ignora pequena diferença de ano e retorna esse.
                return movie
        
        # 2. Se NÃO tem diretor no Excel, vai pelo Ano estrito
        elif check_year_match(movie.get('release_date'), ano):
            return movie

    # --- MUDANÇA CRUCIAL ---
    # Se tinha diretor no Excel e NENHUM bateu, retorna None.
    # Não pegamos mais o "primeiro da lista" por chute.
    if diretor_excel:
        return None
        
    return results[0] if results else None

def save_to_supabase(data):
    url = f"{SUPABASE_URL}/rest/v1/reviews"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal"
    }
    requests.post(url, json=data, headers=headers)

def migrate_data():
    print("--- MIGRAÇÃO RIGOROSA (DIRETOR OBRIGATÓRIO) ---")
    try:
        wb = load_workbook(EXCEL_FILE)
    except:
        print("Excel não encontrado.")
        return

    # Acha a aba
    target_sheet = wb.active
    for name in wb.sheetnames:
        if "ju" in name.lower() or "joão" in name.lower():
            target_sheet = wb[name]
            break

    # Cabeçalhos
    header_row_idx = -1
    headers = []
    for i, row in enumerate(target_sheet.iter_rows(max_row=10, values_only=True), start=1):
        row_str = [str(c).lower() for c in row if c]
        if any("nome" in s for s in row_str) and any(x in str(row_str) for x in ['nota', 'avaliação']):
            header_row_idx = i
            headers = row
            break
            
    if header_row_idx == -1: return

    # Mapeamento
    def get_idx(opts):
        for idx, h in enumerate(headers):
            if h and str(h).lower().strip() in [o.lower() for o in opts]: return idx
        return -1

    idx_nome_pt = get_idx(['Nome pt-BR', 'Nome'])
    idx_nome_orig = get_idx(['Nome original', 'Original'])
    idx_diretor = get_idx(['Diretor', 'Director']) 
    idx_ano = get_idx(['Ano de lançamento', 'Ano'])
    idx_nota = get_idx(['Avaliação 0 -10', 'Nota'])
    idx_review = get_idx(['Principais comentários', 'Comentários'])
    idx_rec = get_idx(['Recomendamos?', 'Recomendação'])

    if idx_nome_pt == -1: idx_nome_pt = 1

    print("\nIniciando...")
    rows = list(target_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True))
    
    count = 0
    for row in rows:
        if not row or len(row) <= idx_nome_pt or not row[idx_nome_pt]: continue

        nome_pt = row[idx_nome_pt]
        nome_orig = row[idx_nome_orig] if idx_nome_orig != -1 else None
        diretor = str(row[idx_diretor]).strip() if idx_diretor != -1 and row[idx_diretor] else None
        
        ano_val = row[idx_ano] if idx_ano != -1 else None
        ano = str(ano_val).strip() if ano_val and str(ano_val).isdigit() else None
        
        nota = row[idx_nota] if idx_nota != -1 else 0
        review = row[idx_review] if idx_review != -1 else ""
        rec_val = row[idx_rec] if idx_rec != -1 else ""
        rec_texto = str(rec_val).strip() if rec_val else "Não avaliado"

        print(f"> {nome_pt} ({ano or '?'})", end=" ")

        # Tenta pelo Original primeiro
        match = None
        if nome_orig:
            match = find_best_match(nome_orig, ano, diretor)
        
        # Se falhar, tenta pelo PT-BR
        if not match:
            match = find_best_match(nome_pt, ano, diretor)

        if match:
            print(f"-> [OK] {match['title']} (ID: {match['id']})")
            data = {
                "tmdb_id": match['id'],
                "rating": float(nota) if isinstance(nota, (int, float)) else 0,
                "review": str(review) if review else "",
                "recommended": rec_texto
            }
            save_to_supabase(data)
            count += 1
        else:
            print("-> [X] NÃO ENCONTRADO (Diretor/Ano não bateram)")

    print(f"\n--- FIM: {count} importados ---")

if __name__ == "__main__":
    migrate_data()